import os
import pathlib
import pandas as pd
import openpyxl

current_dir=os.path.dirname(os.path.abspath(__file__))
work_dir=pathlib.Path(current_dir).parent.parent
target_dir=work_dir.joinpath('result').joinpath('baseline').joinpath('CNT')
folder_list=os.listdir(target_dir)
workbook=openpyxl.Workbook()
workbook_dir=work_dir.joinpath('result').joinpath('baseline')
sheet=workbook.active
col=1
for folder in folder_list:
    row=1
    forder_dir=target_dir.joinpath(folder)
    file_list=os.listdir(forder_dir)
    sheet.cell(row=1, column=col+1, value=str(folder))
    for file in file_list:
        file_dir=forder_dir.joinpath(file)
        data= pd.read_csv(file_dir)
        len_file = len(data)
        sheet.cell(column=1, row=row, value=file)
        for i in range(0,len_file):
            sheet.cell(column=1,row=row+i+1, value = data.iloc[i,1])
        for i in range(0,len_file):
            sheet.cell(column=col+1,row=row+i+1,value=data.iloc[i,2])
        row= row+len_file +1
    col = col +1

workbook.save(workbook_dir.joinpath('CNTbaseline.xlsx'))