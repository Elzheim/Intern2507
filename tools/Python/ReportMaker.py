import os
import pathlib
import pandas as pd
import openpyxl

current_dir=os.path.dirname(os.path.abspath(__file__))
work_dir=pathlib.Path(current_dir).parent.parent
result_dir=work_dir.joinpath('result')
file_list=os.listdir(result_dir)
target=result_dir.joinpath(file_list[2])
folder_list=os.listdir(target)

workbook=openpyxl.Workbook()
workbook_dir=result_dir.joinpath('Summary.xlsx')
workbook.save(workbook_dir)
for folder in folder_list:
    folder_dir=target.joinpath(folder)
    sheet=workbook.create_sheet(folder)
    files=os.listdir(folder_dir)
    file_cnt=len(file_list)
    len_df0=0
    rows=0
    cols=0
    for file in files:
        file_dir=folder_dir.joinpath(file)
        df_file=pd.read_csv(file_dir)
        name_data=file.split(' ')
        model_data=name_data[1].split('.')
        cycle_data=name_data[0]
        model_data=model_data[0]
        len_df1=len(df_file)
        if len_df0>len_df1:
            cols= cols+1
            rows = 0
        sheet.cell(row=rows+1,column=1,value='model '+model_data)
        sheet.cell(row=1, column=cols+2, value=cycle_data)
        for i in range(0,len_df1):
            sheet.cell(row=rows+i+2,column=1,value=df_file.iloc[i,0])
        for i in range(0,len_df1):
            sheet.cell(row=rows+i+2,column=cols+2, value=str(round(df_file.iloc[i,1],4))+"("+str(round(df_file.iloc[i,2],4))+','+str(round(df_file.iloc[i,3],4))+', p value='+str(round(df_file.iloc[i,4],4))+')')
        rows= rows+len_df1+1
        len_df0=len_df1

workbook.save(result_dir.joinpath('Summary.xlsx'))

